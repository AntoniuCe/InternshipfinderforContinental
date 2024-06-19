import requests
import json
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

wsConti = wb.create_sheet("Conti ", 0)
wsNokia = wb.create_sheet("Nokia ", 1)

def jobs():
    payload = {
        'tx_conjobs_api[filter][locationSuggestChecksums][]': '2525a20f75835ba068814ce776b101b2',
        'tx_conjobs_api[filter][searchTerm]': 'Internship',
        'tx_conjobs_api[itemsPerPage]': '100'
    }
    url = "https://jobs.continental.com/ro/api/result-list/pagetype-jobs/"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    }

    response = requests.get(url, params=payload, headers=headers, verify=False)
    json_data = response.json()
    if response.status_code == 200:
        json_data = response.json()
        if "result" in json_data:
            job_list = json_data["result"]

            file_path = "jobs_data.json"

            with open(file_path, "w") as json_file:
                json.dump(job_list, json_file, indent=4)

            print(f"JSON data has been saved to {file_path}")
        else:
            print("No data list found in JSON")
    else:
        print(f"Status code != 200: {response.status_code}")
    with open('jobs_data.json') as f:
        data = json.load(f)
        headers = list(data['list'][0].keys())
        wsConti.append(headers)

    for row_data in data['list']:
        row = [row_data[key] for key in headers]
        wsConti.append(row)



    payload = {
        'onlyData': 'true',
        'expand': 'requisitionList.secondaryLocations,flexFieldsFacet.values',
        'finder': 'findReqs;siteNumber=CX_1,facetsList=LOCATIONS%3BWORK_LOCATIONS%3BWORKPLACE_TYPES%3BTITLES%3BCATEGORIES%3BORGANIZATIONS%3BPOSTING_DATES%3BFLEX_FIELDS,limit=10,keyword=%22Internship%20%22,locationId=100000027583723,radius=25,radiusUnit=MI,sortBy=RELEVANCY'
    }
    url = "https://fa-evmr-saasfaprod1.fa.ocs.oraclecloud.com/hcmRestApi/resources/latest/recruitingCEJobRequisitions?onlyData=true&expand=requisitionList.secondaryLocations,flexFieldsFacet.values&finder=findReqs;siteNumber=CX_1,facetsList=LOCATIONS%3BWORK_LOCATIONS%3BWORKPLACE_TYPES%3BTITLES%3BCATEGORIES%3BORGANIZATIONS%3BPOSTING_DATES%3BFLEX_FIELDS,limit=10,keyword=%22Internship%20%22,locationId=100000027583723,radius=25,radiusUnit=MI,sortBy=RELEVANCY"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    }

    response = requests.get(url, params=payload, headers=headers, verify=False)

    if response.status_code == 200:
        json_data = response.json()

        if "items" in json_data:
            job_list = []
            for item in json_data["items"]:
                if "requisitionList" in item:
                    job_list.extend(item["requisitionList"])

            file_path = "jobs_data_nokia.json"

            with open(file_path, "w") as json_file:
                json.dump(job_list, json_file, indent=4)

            print(f"JSON data has been saved to {file_path}")
        else:
            print("No data list found in JSON")
    else:
        print(f"Status code != 200: {response.status_code}")

    with open('jobs_data_nokia.json') as f:
        nokia_data = json.load(f)
        headers = list(nokia_data[0].keys())
        wsNokia.append(headers)
    for row_data in nokia_data:
        wsNokia.append(row)


jobs()
wb.save("jobs.xlsx")