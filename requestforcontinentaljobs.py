import requests
import json
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

wsConti = wb.create_sheet("Conti ", 0)
wsNokia = wb.create_sheet("Nokia ", 1)
wsZF = wb.create_sheet("ZF ", 2)

'''wsConti = wb["Continental jobs"]
wsNokia = wb["Nokia jobs"]
wsZF = wb["ZF jobs"]'''


def conti():
    payload = {
        'tx_conjobs_api[filter][locationSuggestChecksums][]': '2525a20f75835ba068814ce776b101b2',
        'tx_conjobs_api[filter][searchTerm]': 'Internship',
        'tx_conjobs_api[itemsPerPage]': '100'
    }
    url = "https://jobs.continental.com/ro/api/result-list/pagetype-jobs/"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    }

    response = requests.get(url, params=payload, headers=headers, verify=False)
    json_data = response.json()

    if "result" in json_data:
        job_list = json_data["result"]

        file_path = "jobs_data.json"

        with open(file_path, "w") as json_file:
            json.dump(job_list, json_file, indent=4)

        print(f"datele json s-au salvat in {file_path}")
    else:
        print("nu s-o gasit nicio lista de date in json")

    with open('jobs_data.json') as f:
        data = json.load(f)

        headers = list(data['list'][0].keys())
        wsConti.append(headers)
    for row_data in data['list']:
        row = [row_data[key] for key in headers]
        wsConti.append(row)



conti()
wb.save("jobs.xlsx")
